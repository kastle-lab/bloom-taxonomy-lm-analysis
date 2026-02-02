from itertools import combinations
import torch
from bert_score import score
from openpyxl import load_workbook
import os
import warnings
warnings.filterwarnings("ignore")

os.environ["TRANSFORMERS_NO_ADVISORY_WARNINGS"] = "true"


# =====================
# CONFIG
# =====================
INPUT_FILE = "Large_Model_Questions.xlsx"
OUTPUT_FILE = "Large_Model_Questions_WITH_BERTSCORE_PRF.xlsx"

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
# MODEL_TYPE = "bert-base-uncased"
MODEL_TYPE = "microsoft/deberta-xlarge-mnli"

# =====================
# LOAD WORKBOOK
# =====================
wb = load_workbook(INPUT_FILE)

# =====================
# PROCESS SHEETS (EXCEPT LAST)
# =====================
for sheet_name in wb.sheetnames[:-1]:
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]

    # Robust header lookup (handles whitespace / merges)
    def find_col(name):
        for idx, h in enumerate(headers):
            if isinstance(h, str) and h.strip().lower() == name.lower():
                return idx + 1
        raise ValueError(f"Column '{name}' not found in sheet '{sheet_name}'")

    model_col = find_col("Model")
    question_col = find_col("Questions")

    # =====================
    # CREATE OUTPUT COLUMNS
    # =====================
    start_col = len(headers) + 1
    pairs = list(combinations(range(1, 7), 2))  # (1,2)...(5,6)

    col_map = {}
    col_ptr = start_col

    for i, j in pairs:
        for metric in ["p", "r", "f1"]:
            header = f"sim_{i}_{j}_{metric}"
            ws.cell(row=1, column=col_ptr).value = header
            col_map[(i, j, metric)] = col_ptr
            col_ptr += 1

    # =====================
    # MAIN LOOP
    # =====================
    row = 2
    max_row = ws.max_row
    current_model = None

    while row <= max_row:
        # Handle merged model cells
        cell_value = ws.cell(row, model_col).value
        if cell_value is not None:
            current_model = cell_value

        if current_model is None:
            row += 1
            continue

        model_name = current_model

        # Collect contiguous rows for this model
        block_rows = []
        r = row
        temp_model = current_model

        while r <= max_row:
            val = ws.cell(r, model_col).value
            if val is not None:
                temp_model = val
            if temp_model != model_name:
                break
            block_rows.append(r)
            r += 1

        # =====================
        # PROCESS ONLY 6-ROW BLOCKS
        # =====================
        if len(block_rows) == 6:
            questions = [
                str(ws.cell(rr, question_col).value) for rr in block_rows
            ]

            candidates, references, pair_order = [], [], []

            for i, j in pairs:
                references.append(questions[i - 1])
                candidates.append(questions[j - 1])
                pair_order.append((i, j))

            P, R, F1 = score(
                candidates,
                references,
                model_type=MODEL_TYPE,
                device=DEVICE,
                lang="en",
                verbose=False
            )

            # Write + merge output
            for idx, (i, j) in enumerate(pair_order):
                for metric, tensor in zip(["p", "r", "f1"], [P, R, F1]):
                    col = col_map[(i, j, metric)]
                    ws.cell(row=row, column=col).value = float(tensor[idx])
                    ws.merge_cells(
                        start_row=row,
                        end_row=row + 5,
                        start_column=col,
                        end_column=col
                    )

        # Move to next block
        row = r

# =====================
# SAVE
# =====================
wb.save(OUTPUT_FILE)
print("BERTScore Precision / Recall / F1 written successfully with merged formatting preserved.")
