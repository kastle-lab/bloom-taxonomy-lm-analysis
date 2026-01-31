import warnings
warnings.filterwarnings("ignore")

import os
os.environ["TRANSFORMERS_NO_ADVISORY_WARNINGS"] = "true"

from openpyxl import load_workbook
from bert_score import score
import torch
from itertools import combinations

INPUT_FILE = "grouped_output_merged_ALLSHEETS-BloomAndGradeLevelsCheck.xlsx"
OUTPUT_FILE = "grouped_output_merged_ALLSHEETS-BloomAndGradeLevelsCheck_WITH_BERTSCORE_PRF.xlsx"

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
# MODEL_TYPE = "bert-base-uncased"
MODEL_TYPE = "microsoft/deberta-xlarge-mnli"

wb = load_workbook(INPUT_FILE)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]

    q_col = headers.index("Questions") + 1
    bloom_col = headers.index("Bloom's level") + 1

    # Start writing at first empty column
    start_col = len(headers) + 1

    pairs = list(combinations(range(1, 7), 2))

    # Build column map: (i,j,metric) -> column index
    col_map = {}
    col_ptr = start_col

    for i, j in pairs:
        for metric in ["p", "r", "f1"]:
            header = f"sim_{i}_{j}_{metric}"
            ws.cell(row=1, column=col_ptr).value = header
            col_map[(i, j, metric)] = col_ptr
            col_ptr += 1

    row = 2
    max_row = ws.max_row

    while row <= max_row:
        if ws.cell(row, bloom_col).value == 1:
            block_rows = list(range(row, row + 6))
            questions = [
                str(ws.cell(r, q_col).value) for r in block_rows
            ]

            # Build batched candidate / reference lists
            candidates = []
            references = []
            pair_order = []

            for i, j in pairs:
                references.append(questions[i - 1])
                candidates.append(questions[j - 1])
                pair_order.append((i, j))

            P, R, F1 = score(
                candidates,
                references,
                model_type=MODEL_TYPE,
                device=DEVICE,
                lang="en",
                verbose=False
            )

            for idx, (i, j) in enumerate(pair_order):
                for metric, tensor in zip(
                    ["p", "r", "f1"], [P, R, F1]
                ):
                    col = col_map[(i, j, metric)]
                    ws.cell(row=row, column=col).value = float(tensor[idx])

                    ws.merge_cells(
                        start_row=row,
                        end_row=row + 5,
                        start_column=col,
                        end_column=col
                    )

            row += 6
        else:
            row += 1

wb.save(OUTPUT_FILE)

print("BERTScore Precision / Recall / F1 written successfully with merged formatting preserved.")

